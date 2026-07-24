#!/usr/bin/env python3
"""
Knowledge translation pipeline: CS TikTok Shop Excel -> Open WebUI knowledge markdown.

Usage: python3 generate_kb.py <input.xlsx> <output_dir>

Sheets consumed:
  - "Template for KnowledgeBase"      (canonical FAQ; superset)
  - "Categorized issues_FAQ provided" (merged in: EN translations + routing flags)
  - "Common Violations"
  - "Q&A Creator Policy Q12025"
  - "Terminology"

Output: one .md per FAQ sub-category, plus violations / policy / terminology /
escalation-guide files. Thai-primary with bilingual headers. Link-only answers
are kept as official-link references (Phase 1: no TikTok Academy scraping).

Every generated content file (not the directory READMEs) carries governance
frontmatter (id/title/audiences/owner/sensitivity/status/source_refs/
reviewed_at/review_by) per the Content Contract in
ai/architecture/llm-knowledge-base-plan.md. This is metadata only, not yet
enforced: nothing here or in upload_kb.py validates it or derives Open WebUI
access grants from it -- the collection's real grants are set and maintained
manually and remain the actual source of truth regardless of what this
frontmatter says (see ai/openwebui/knowledge/creator-services/README.md
"Governance status"). This collection maps to the plan's roadmapped
`wiki-erisa` slot (Erisa groups / creator & affiliate workflows), reached
early via this lighter bootstrap pipeline instead of the full company-wiki
validator/Sync Pipe.

Repo home: scripts/ai/creator-kb/generate_kb.py, with output synced to
ai/openwebui/knowledge/creator-services/.
"""
import os
import re
import shutil
import sys
import unicodedata
from collections import OrderedDict

from openpyxl import load_workbook

LINK_RE = re.compile(r"https?://\S+")

GOVERNANCE = {
    "owner": "erisa-creator-services",
    # Must be a value from company-wiki/tools/wiki-schema.json's audiences
    # enum (erisa-member/erisa-team-lead/erisa-manager, or the "erisa"
    # shorthand) so this metadata is reusable by that validator later.
    "audiences": ["erisa"],
    "sensitivity": "department",
    "source_ref": "CS_TikTok_Shop__Knowledge_Base.xlsx",
    # Frozen to the date an owner actually reviewed this classification, not
    # datetime.date.today(): a mechanical regeneration is not a review, and
    # advancing these on every run would falsely certify stale content
    # (e.g. Q1/2025 policy pages) as freshly reviewed. Bump both by hand only
    # after an actual owner review.
    "reviewed_at": "2026-07-23",
    "review_by": "2026-10-21",
}


def frontmatter(id_, title):
    return [
        "---",
        f"id: {id_}",
        f"title: {title}",
        f"audiences: [{', '.join(GOVERNANCE['audiences'])}]",
        f"owner: {GOVERNANCE['owner']}",
        f"sensitivity: {GOVERNANCE['sensitivity']}",
        "status: active",
        f"source_refs: [{GOVERNANCE['source_ref']}]",
        f"reviewed_at: {GOVERNANCE['reviewed_at']}",
        f"review_by: {GOVERNANCE['review_by']}",
        "---",
        "",
    ]


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("", "None", "#N/A", "#REF!", "#ERROR!") else s


def qnum(q):
    m = re.match(r"^(\d+\.\d+\.\d+)", q)
    return m.group(1) if m else None


def strip_qnum(q):
    return re.sub(r"^\d+\.\d+\.\d+\s*", "", q).strip()


def slugify(s, maxlen=50):
    s = re.sub(r"\(.*?\)", "", s)
    s = unicodedata.normalize("NFKD", s)
    s = re.sub(r"[^a-zA-Z0-9\s-]", "", s).strip().lower()
    s = re.sub(r"[\s_]+", "-", s)
    return s[:maxlen].rstrip("-") or "misc"


def rows_of(wb, name):
    return [[clean(c) for c in row] for row in wb[name].iter_rows(values_only=True)]


def parse(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)

    # --- Template for KnowledgeBase (header at row index 1) ---
    tpl = rows_of(wb, "Template for KnowledgeBase")
    template_records, last_main, last_sub = [], "", ""
    for r in tpl[2:]:
        r = r + [""] * (11 - len(r))
        main_cat, sub_cat, q, a = r[2], r[4], r[5], r[6]
        tags, alt, src, human = r[7], r[8], r[9], r[10]
        if main_cat:
            last_main = main_cat
        if sub_cat:
            last_sub = sub_cat
        if not q and not a:
            continue
        template_records.append({
            "main": last_main, "sub": last_sub, "question": q, "answer": a,
            "tags": tags, "keywords": alt, "source": src,
            "needs_human": human.lower() in ("yes", "y", "true"),
        })

    # --- Categorized issues (2 header rows, merged cells filled down) ---
    cat = rows_of(wb, "Categorized issues_FAQ provided")
    cat_by_num = {}
    for r in cat[2:]:
        r = r + [""] * (11 - len(r))
        q, q_en = r[2], r[3]
        edu, selfhelp, human_am, platform = r[7], r[8], r[9], r[10]
        if not q:
            continue
        if q_en.startswith("=IFERROR"):
            q_en = ""
        routes = [name for flag, name in (
            (edu, "education"), (selfhelp, "self_help"),
            (human_am, "human_am"), (platform, "platform_ticket")) if flag]
        n = qnum(q)
        if n:
            cat_by_num[n] = {"question_en": strip_qnum(q_en) if q_en else "", "routes": routes}

    # Merge EN + routes into template records; classify answer type
    for r in template_records:
        extra = cat_by_num.get(qnum(r["question"]) or "", {})
        r["question_en"] = extra.get("question_en", "")
        r["routes"] = extra.get("routes", [])
        a = r["answer"]
        if "Request Admin Help" in a or r["needs_human"]:
            r["type"] = "escalation"
        elif a.startswith("ศึกษารายละเอียด") and LINK_RE.search(a):
            r["type"] = "link_ref"
        else:
            r["type"] = "full"

    # --- Common Violations ---
    vio_records = []
    for r in rows_of(wb, "Common Violations")[1:]:
        r = r + [""] * (13 - len(r))
        name_en, name_th = r[3], r[4]
        if name_en.startswith("=IFERROR"):
            name_en = ""
        if not name_th:
            continue
        vio_records.append({
            "name_th": name_th, "name_en": name_en, "keywords": r[5],
            "description": r[6], "review": r[7], "evidence": r[8],
            "learn_more": r[9], "prevention": r[10], "penalty": r[11], "at_risk": r[12],
        })

    # --- Q&A Creator Policy ---
    pol_records = []
    for r in rows_of(wb, "Q&A Creator Policy Q12025")[1:]:
        r = r + [""] * (4 - len(r))
        if not r[1]:
            continue
        pol_records.append({"category": r[0], "question": r[1], "answer": r[2], "note": r[3]})

    # --- Terminology ---
    term_records = []
    for r in rows_of(wb, "Terminology")[1:]:
        r = r + [""] * (6 - len(r))
        if not r[0] and not r[5]:
            continue
        term_records.append({"term_en": r[0], "category": r[1], "aliases": r[2],
                             "not_confused_with": r[3], "definition_en": r[4],
                             "definition_th": r[5]})

    return template_records, vio_records, pol_records, term_records


def fmt_answer(r):
    if r["type"] == "escalation":
        return ("**การดำเนินการ:** เรื่องนี้ต้องส่งต่อให้แอดมิน/ทีม Creator Services ดูแลโดยตรง "
                "ไม่สามารถแก้ไขด้วยตนเองได้\n"
                "**Action:** This issue requires admin / Creator Services escalation. "
                "Do not attempt self-service resolution.")
    if r["type"] == "link_ref":
        url = LINK_RE.search(r["answer"]).group(0)
        return (f"ศึกษารายละเอียดจากคู่มืออย่างเป็นทางการของ TikTok Shop Academy:\n{url}\n\n"
                f"_(หมายเหตุสำหรับผู้ช่วย AI: เนื้อหาบทความนี้ยังไม่ได้นำเข้าฐานความรู้ "
                f"ให้ส่งลิงก์นี้ให้ครีเอเตอร์โดยตรง)_")
    return r["answer"]


GENERATED_SUBDIRS = ("faq", "policy", "violations", "terminology")
ESCALATION_GUIDE_NAME = "00-escalation-guide.md"


def generate(records, vio, pol, term, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    # Clear prior-run output before writing so a removed/renamed Excel
    # category doesn't leave its old .md file behind as still-publishable
    # source content. README.md is hand-maintained and never touched here.
    for sub in GENERATED_SUBDIRS:
        shutil.rmtree(os.path.join(out_dir, sub), ignore_errors=True)
    esc_path = os.path.join(out_dir, ESCALATION_GUIDE_NAME)
    if os.path.exists(esc_path):
        os.remove(esc_path)
    written = []

    # --- FAQ files, one per (main, sub) ---
    groups = OrderedDict()
    for r in records:
        groups.setdefault((r["main"], r["sub"]), []).append(r)
    faq_dir = os.path.join(out_dir, "faq")
    os.makedirs(faq_dir, exist_ok=True)
    for gi, ((main, sub), items) in enumerate(groups.items(), start=1):
        slug_id = f"{slugify(main, 30)}--{slugify(sub, 40)}"
        fname = f"{gi:02d}-{slug_id}.md"
        title = f"{main} — {sub}"
        lines = frontmatter(f"creator-services.faq.{slug_id}", title) + [
            f"# {main} — {sub}", "",
            f"> ฐานความรู้สำหรับผู้ช่วย Creator Services (ERISA) | หมวด: {main} / {sub}",
            f"> Knowledge base for the ERISA creator-service assistant. Category: {main} / {sub}.",
            "> Audience: TikTok Shop creators (Thailand). Source: CS TikTok Shop Knowledge Base.",
            "",
        ]
        for r in items:
            n = qnum(r["question"]) or ""
            lines.append(f"## Q{(' ' + n) if n else ''}: {strip_qnum(r['question'])}")
            if r.get("question_en"):
                lines.append(f"_EN: {r['question_en']}_")
            if r.get("keywords"):
                lines.append(f"**คำค้นที่เกี่ยวข้อง / Keywords:** {r['keywords']}")
            lines += ["", fmt_answer(r)]
            if r.get("source") and r["type"] != "link_ref":
                lines.append(f"\n**อ้างอิง / Source:** {r['source']}")
            if r["type"] == "escalation":
                lines.append("\n**Escalation:** ✋ ส่งต่อแอดมิน (Request Admin Help)")
            lines.append("\n---\n")
        with open(os.path.join(faq_dir, fname), "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        written.append(("faq/" + fname, len(items)))

    # --- Violations ---
    vl = frontmatter("creator-services.violations.common-violations",
                      "Common Creator Violations") + [
        "# การละเมิดที่พบบ่อยของครีเอเตอร์ TikTok Shop — Common Creator Violations", "",
        "> ฐานความรู้: ประเภทการละเมิด หลักฐานที่ต้องใช้ในการอุทธรณ์ วิธีป้องกัน และบทลงโทษ",
        "> Violation types, appeal evidence requirements, prevention guidance, and penalties.", "",
    ]
    for v in vio:
        vl.append(f"## {v['name_th']}" + (f" ({v['name_en']})" if v["name_en"] else ""))
        for key, label in (("keywords", "**Keywords:**"),
                           ("description", "\n**คำอธิบาย:**"),
                           ("review", "\n**แนวทางการตรวจสอบ:**"),
                           ("evidence", "\n**หลักฐานที่ต้องใช้ในการอุทธรณ์:**"),
                           ("prevention", "\n**วิธีป้องกัน:**"),
                           ("penalty", "\n**บทลงโทษ:**"),
                           ("at_risk", "\n**กลุ่มครีเอเตอร์ที่เสี่ยงที่สุด:**"),
                           ("learn_more", "\n**เรียนรู้เพิ่มเติม:**")):
            if v[key]:
                vl.append(f"{label} {v[key]}")
        vl.append("\n---\n")
    vdir = os.path.join(out_dir, "violations")
    os.makedirs(vdir, exist_ok=True)
    with open(os.path.join(vdir, "common-violations.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(vl))
    written.append(("violations/common-violations.md", len(vio)))

    # --- Policy Q&A grouped by category ---
    pol_groups = OrderedDict()
    for p in pol:
        pol_groups.setdefault(p["category"] or "อื่น ๆ", []).append(p)
    pdir = os.path.join(out_dir, "policy")
    os.makedirs(pdir, exist_ok=True)
    for pi, (cat, items) in enumerate(pol_groups.items(), start=1):
        slug = slugify(cat, 40)
        slug_id = slug if (slug != "misc" and re.search(r"[a-z]", slug)) else "policy"
        fname = f"{pi:02d}-{slug_id}.md"
        lines = frontmatter(f"creator-services.policy.{slug_id}",
                             f"Creator Policy Q1/2025 — {cat}") + [
            f"# นโยบายครีเอเตอร์ Q1/2025 — {cat}", "",
            f"> Q&A นโยบายครีเอเตอร์ TikTok Shop ไตรมาส 1/2025 หมวด: {cat}",
            f"> TikTok Shop creator policy Q&A (Q1/2025). Category: {cat}", "",
        ]
        for p in items:
            lines += [f"## Q: {p['question']}", "",
                      p["answer"] or "_(ยังไม่มีแนวทางคำตอบ — ส่งต่อแอดมิน / escalate to admin)_"]
            if p["note"]:
                lines.append(f"\n_Note: {p['note']}_")
            lines.append("\n---\n")
        with open(os.path.join(pdir, fname), "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        written.append(("policy/" + fname, len(items)))

    # --- Terminology ---
    tdir = os.path.join(out_dir, "terminology")
    os.makedirs(tdir, exist_ok=True)
    tl = frontmatter("creator-services.terminology.glossary",
                      "Terminology Glossary") + [
        "# อภิธานศัพท์ TikTok Shop / ERISA — Terminology Glossary", "",
        "> คำศัพท์เฉพาะทางที่ใช้ในงาน Creator Services และ TikTok Shop",
        "> Terms used in creator services and TikTok Shop operations.", "",
    ]
    for t in term:
        tl.append(f"## {t['term_en'] or t['definition_th'][:40]}")
        for key, label in (("category", "**หมวด / Category:**"),
                           ("aliases", "**ชื่อที่ใช้เรียกทั่วไป / Also known as:**"),
                           ("not_confused_with", "**อย่าสับสนกับ / Not to be confused with:**"),
                           ("definition_en", "\n**Definition (EN):**"),
                           ("definition_th", "\n**คำจำกัดความ (TH):**")):
            if t[key]:
                tl.append(f"{label} {t[key]}")
        tl.append("\n---\n")
    with open(os.path.join(tdir, "glossary.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(tl))
    written.append(("terminology/glossary.md", len(term)))

    # --- Escalation guide ---
    esc = [r for r in records if r["type"] == "escalation"]
    el = frontmatter("creator-services.escalation-guide", "Escalation Guide") + [
        "# แนวทางการส่งต่อเรื่องให้แอดมิน — Escalation Guide", "",
        "> กติกาสำหรับผู้ช่วย AI: หัวข้อต่อไปนี้ **ห้ามตอบเอง** ต้องส่งต่อให้แอดมิน/ทีม Creator Services เสมอ",
        "> Rules for the AI assistant: the following topics MUST be escalated to admin / Creator Services.",
        "", "## หัวข้อที่ต้องส่งต่อแอดมินเสมอ / Always-escalate topics", "",
    ]
    for r in esc:
        el.append(f"- **{strip_qnum(r['question'])}** ({r['main']} / {r['sub']})")
    el += [
        "", "## วิธีตอบเมื่อต้องส่งต่อ / How to respond when escalating", "",
        "เมื่อครีเอเตอร์ถามเรื่องในรายการข้างต้น ให้ตอบว่า:",
        "\"เรื่องนี้ทีมงานต้องตรวจสอบให้เป็นรายกรณีค่ะ รบกวนส่งรายละเอียด "
        "(ชื่อบัญชี, ภาพหน้าจอ, วันที่เกิดเหตุ) มาให้แอดมิน "
        "แล้วทีม Creator Services จะติดต่อกลับค่ะ\"", "",
        "Additional escalation triggers beyond this list: account permanently banned, "
        "payment missing beyond 10 business days, suspected system error, "
        "trademark/IP violations, and any request the knowledge base cannot answer confidently.",
    ]
    with open(os.path.join(out_dir, "00-escalation-guide.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(el))
    written.append(("00-escalation-guide.md", len(esc)))
    return written


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    records, vio, pol, term = parse(sys.argv[1])
    written = generate(records, vio, pol, term, sys.argv[2])
    total = sum(n for _, n in written)
    print(f"Generated {len(written)} files, {total} entries -> {sys.argv[2]}")
    for fp, n in written:
        print(f"  [{n:3d}] {fp}")


if __name__ == "__main__":
    main()
